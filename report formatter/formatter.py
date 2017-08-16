import os
import sys
import time
import csv
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill,\
                            Font, GradientFill, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
    


    
# this useful function taken from konrad
# at "https://gist.github.com/konrad/4154786"
def _convert_to_number(cell):
    if cell.isnumeric():
        return int(cell)
    try:
        return float(cell)
    except ValueError:
        return cell
        
        
        
        
def csv_to_xlsx(input_path):
    wb = Workbook()
    worksheet = wb.active
    with open(input_path) as csv_file:
        for row in csv.reader(csv_file, delimiter=","):
            worksheet.append([_convert_to_number(cell) for cell in row])
        
    return (wb, worksheet)
    
    
    
    
def set_print_properties(ws, first_cell, last_cell):
    #setting up the print settings
    ws.oddHeader.center.text = "&F"
    ws.oddHeader.center.size = 14
    ws.oddHeader.center.font = "Tahoma,Bold"
    
    ws.oddFooter.right.text = "Page &[Page] of &N"
    ws.oddFooter.right.size = 12
    ws.oddFooter.right.font = "Tahoma"
    
    ws.oddFooter.left.text = "&[{}".format(time.strftime("%m/%d/%Y")) # THIS IS SOME FReAKING MAGIC JESUS BUT IT WORKS SCREW YOUR FARMATTING OPENPYXL NOR YOUR STUPID &D !!:(
    ws.oddFooter.left.size = 12
    ws.oddFooter.left.font = "Tahoma"
    
    ws.print_options.horizontalCentered = True
    ws.print_area = "{}:{}".format(first_cell.coordinate,
                                    last_cell.coordinate)

    
    
    
# Taken from documentation   
def style_range(ws, first_cell, last_cell):
    # colouring rows after first row
    for r in range(first_cell.row, last_cell.row+1):
        for c in range(first_cell.col_idx, last_cell.col_idx+1): 
            cell = ws.cell(row = r, column = c)
            # colour odd numbered rows
            if r == first_cell.row:
                ft = Font(name='Tahoma', bold=label_bold.get())
                if isinstance(cell.value, str):
                    al = Alignment(wrapText=label_wrap_text.get(), horizontal=label_txtHorAl.get(), vertical=label_vertAl.get())
                else:
                    al = Alignment(wrapText=True, horizontal=label_numHortAl.get(), vertical=label_vertAl.get())
            else:
                ft = Font(name='Tahoma', bold=data_bold.get())
                if isinstance(cell.value, str):
                    al = Alignment(wrapText=data_wrap_text.get(), horizontal=data_txtHorAl.get(), vertical=data_vertAl.get())
                else:
                    al = Alignment(wrapText=True, horizontal=data_numHortAl.get(), vertical=data_vertAl.get())
            border = Border(left=Side(border_style="thin",
                            color='FF000000'),
                    right=Side(border_style="thin",
                            color='FF000000'),
                    top=Side(border_style="thin",
                            color='FF000000'),
                    bottom=Side(border_style="thin"))
            
            cell.alignment = al
            cell.font = ft
            cell.border = border
            if r % 2 == 1:
                fill = PatternFill('solid', fgColor="DDDDDD")
                cell.fill = fill
                
    # fit text to column
    column_widths = []
    for row in ws:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column].width = column_width

        
        
           
def format():
    # input csv file location data
    file_base_name = os.path.basename(input_file_path)
    
    # output xlsx file location data
    output_file_path = "{}_formatted.xlsx".format(input_file_path[:-4])
    
    # open and convert the csv to xlsx 
    try:
        wb, ws = csv_to_xlsx(input_file_path)
    except FileNotFoundError:
        messagebox.showinfo("File not found", "Please use Browse to select a valid .CSV file before attempting to format.")
        return
    
    # the actual formatting
    first_cell = ws.cell(row=1, column=1)
    last_cell = ws.cell(row=ws.max_row, column=ws.max_column)
    style_range(ws, first_cell, last_cell)
    
    set_print_properties(ws, first_cell, last_cell)
    
    # save formatted file
    wb.save(output_file_path)



input_file_path = ""
def load_file():
    global input_file_path
    input_file_path = filedialog.askopenfilename(title='Choose a file', filetypes = [("CSV Files",".csv")])
    pathlabel.config(text=input_file_path)
    
 
showsettings = False
def hide_settings(event):
    global showsettings
    if showsettings is False:
        bottomFrame.grid()
        showsettings = True
        event.widget.config(text="Hide Settings")
    else:
        showsettings = False
        bottomFrame.grid_remove()
        event.widget.config(text="Show Settings")
    
    
    
if __name__ == "__main__":
    root = Tk()
    root.title("CSV REDCap report formatter")
    root.resizable(width=False, height=False)
    
    #top frame
    topFrame = Frame(root)
    topFrame.grid(row = 0)
    browse_button = Button(topFrame, text="Browse", font = "Verdana 12 bold", command=load_file)
    pathlabel = Label(topFrame, font="Verdana 12", width=100)
    format_button = Button(topFrame, text='Format', font = "Verdana 12 bold", command=format)
    settings_button = Button(topFrame, text="Show Settings", font = "Verdana 8", width=12)
    settings_button.bind('<Button-1>', hide_settings)
    browse_button.grid(row = 0, sticky = W, padx=5, pady=5)
    pathlabel.grid(row=0, column = 1, sticky = W, padx=5, pady=5)
    format_button.grid(row=0, column = 2, sticky = W, padx=5, pady=5, rowspan = 2)
    settings_button.grid(row=1, sticky = W, padx=5, pady=5)
    
    #bottom frame
    bottomFrame = Frame(root)
    bottomFrame.grid(row = 1)
    
    #bottom left frame
    blFrame = Frame(bottomFrame)
    blFrame.grid(row = 0, column = 0)
    blframe_label = Label(blFrame, text = "Cell Format Settings", font="Verdana 12")
    boldL = Label(blFrame, text="Bold")
    wraptextL = Label(blFrame, text="Wrap Text")
    txtvertalL = Label(blFrame, text="Cell Vertical Alignment")
    numvertalL = Label(blFrame, text="Textual Cell Horizontal Alignment")
    horalL = Label(blFrame, text="Numerical Horizontal Alignment")
    blframe_label.grid(row = 0, padx=20, pady=5)
    boldL.grid(row = 1,sticky = E, padx=5, pady=5)
    wraptextL.grid(row=2,sticky = E, padx=5, pady=5)
    txtvertalL.grid(row=3,sticky = E, padx=5, pady=5)
    numvertalL.grid(row=4,sticky = E, padx=5, pady=5)
    horalL.grid(row=5,sticky = E, padx=5, pady=5)
    
    
    #bottom middle frame
    bmFrame = Frame(bottomFrame)
    data_bold=IntVar()
    data_wrap_text=IntVar()
    data_wrap_text.set(1)
    data_vertAl= StringVar(bmFrame)
    data_vertAl.set("center")
    data_numHortAl= StringVar(bmFrame)
    data_numHortAl.set("right")
    data_txtHorAl= StringVar(bmFrame)
    data_txtHorAl.set("center")
    bmFrame.grid(row = 0, column = 1)
    bmframe_label = Label(bmFrame, text = "Data Formatting", font="Verdana 12")
    bmframe_label.grid(row = 0, padx=20, pady=5)
    dboldB = Checkbutton(bmFrame, variable=data_bold)
    dwraptxtB = Checkbutton(bmFrame, variable=data_wrap_text)
    dvertalB = OptionMenu(bmFrame, data_vertAl,"center", "top", "bottom")
    dnumhoralB = OptionMenu(bmFrame, data_numHortAl, "center", "right", "left")
    dtxthoralB = OptionMenu(bmFrame, data_txtHorAl, "center", "right", "left")
    dboldB.grid(row=1)
    dwraptxtB.grid(row=2)
    dvertalB.grid(row=3)
    dnumhoralB.grid(row=5)
    dtxthoralB.grid(row=4)
    
    
    
    #bottom right frame
    brFrame = Frame(bottomFrame)
    label_bold=IntVar()
    label_bold.set(1)
    label_wrap_text=IntVar()
    label_wrap_text.set(1)
    label_vertAl= StringVar(brFrame)
    label_vertAl.set("center")
    label_numHortAl= StringVar(brFrame)
    label_numHortAl.set("right")
    label_txtHorAl= StringVar(brFrame)
    label_txtHorAl.set("center")
    brFrame.grid(row = 0, column = 2)
    brframe_label = Label(brFrame, text = "Label Formatting", font="Verdana 12")
    brframe_label.grid(row = 0, padx=20, pady=5)
    lboldB = Checkbutton(brFrame, variable=label_bold)
    lwraptxtB = Checkbutton(brFrame, variable=label_wrap_text)
    lvertalB = OptionMenu(brFrame, label_vertAl,"center", "top", "bottom")
    lnumhoralB = OptionMenu(brFrame, label_numHortAl, "center", "right", "left")
    ltxthoralB = OptionMenu(brFrame, label_txtHorAl, "center", "right", "left")
    lboldB.grid(row=1)
    lwraptxtB.grid(row=2)
    lvertalB.grid(row=3)
    lnumhoralB.grid(row=5)
    ltxthoralB.grid(row=4)
    
    bottomFrame.grid_remove()
    root.mainloop()